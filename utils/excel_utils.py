# excel_utils.py

from typing import Any, Dict, List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

def style_header(row):
    header_font = Font(bold=True, color="000000")
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    for cell in row:
        cell.font = header_font
        cell.fill = header_fill

def autosize_columns(ws) -> None:
    """Auto-resize columns based on the widest cell in each column."""
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max_len + 2

def write_sheet(ws, data: List[Dict[str, Any]], mapping: Dict[str, str]) -> None:
    """Write API data into a sheet with given header-to-field mapping."""
    headers = list(mapping.keys())
    ws.append(headers)
    style_header(ws[1])

    for item in data:
        row = [item.get(field) for field in mapping.values()]
        ws.append(row)

    autosize_columns(ws)

def create_excel(users, posts, todos, filename: str = "daily_report.xlsx") -> str:
    """Create a multi-sheet Excel workbook and save to filename."""
    wb = Workbook()

    # Sheet 1 — Users
    ws_users = wb.active
    ws_users.title = "Users"
    write_sheet(ws_users, users, {
        "ID": "id",
        "Name": "name",
        "Username": "username",
        "Email": "email",
    })

    # Sheet 2 — Posts
    ws_posts = wb.create_sheet("Posts")
    write_sheet(ws_posts, posts, {
        "ID": "id",
        "User ID": "userId",
        "Title": "title",
    })

    # Sheet 3 — Todos
    ws_todos = wb.create_sheet("Todos")
    write_sheet(ws_todos, todos, {
        "ID": "id",
        "User ID": "userId",
        "Title": "title",
        "Completed": "completed",
    })

    wb.save(filename)
    return filename
